"""`vin` addon - LWIN (Liv-ex Wine Identification Number), offline, no key.

Tools (descriptions intentionally in French - they are runtime UX for
French-speaking agents, see README):
  - vin_recherche : free text (+ country / colour / vintage) -> LWIN7 candidates
  - vin_lwin      : decode a LWIN7 / 11 / 16 / 18 and resolve it against the base
  - vin_code      : compose a LWIN11 / 16 / 18 from a LWIN7 + vintage + format
  - vin_base      : state of the local base (source, date, rows, rejects)

Why this addon exists: an EAN is a poor key for a cellar. Many growers carry no
barcode at all, and those who do routinely reuse the SAME code across vintages -
so a scan identifies "Chateau X, 75 cl" and loses the one attribute a cellar is
built on. LWIN is the industry's actual identifier, it is hierarchical (wine ->
vintage -> format -> pack), and Liv-ex publishes the whole base under Creative
Commons. It is the spine; a barcode is at best a shortcut into it.

**There is no public LWIN API.** `api.liv-ex.com/lwin/search/v1/lwinSearch` is
what the LWIN website itself calls, and it answers `401 Unauthorized` without a
Liv-ex account token (measured 2026-08-11). So this addon never phones home: it
reads the downloaded file, and that is the entire network surface - zero. The
file is free but sits behind a free Liv-ex registration, i.e. a **manual gesture
by the owner**, like the kubeconfig: download from liv-ex.com/lwin/, drop it on
the volume, point `LWIN_DB` at it. Absent, the addon mounts `degraded` and every
tool says so in French rather than failing.

The traps, each verified against the Liv-ex LWIN Guide v1.2 or a published
worked example rather than assumed - the first three fail SILENTLY:

1. **LWIN carries no diacritics.** The guide is explicit: e, o/, a°, c,, n~, ss
   are not supported, and "Spatlese" is stored for "Spätlese". A French user
   types "Château Léoville" and a naive `LIKE` finds nothing, on a base that
   holds the wine. Both sides are therefore folded - and the fold has to cover
   the letters Unicode does NOT decompose (o/, ss, ae, oe, d-, l-), which NFKD
   alone leaves standing.

2. **LWIN18 is not LWIN16 plus two digits.** The pack size is inserted in the
   MIDDLE, between vintage and bottle size. Liv-ex's own example, Leoville
   Barton 2009: 75 cl -> `1012361200900750` (7+4+5), 12x75 cl ->
   `1012361 2009 12 00750` (7+4+2+5). Append the pack at the end and you mint a
   syntactically perfect code that names a different bottle.

3. **Bottle size is five digits of millilitres, zero-padded** (`00750`), and the
   vintage is four digits. Any code that has been through an integer at some
   point has lost its leading zeros and is silently wrong. Same disease upstream:
   an Excel-to-CSV export strips the leading zero of a 7-digit LWIN, so codes
   shorter than 7 are re-padded here rather than rejected.
   Third-party pages describe that field as centilitres; Liv-ex's own worked
   example is millilitres, and the example wins.

4. A composed LWIN11 is structurally valid, which is NOT the same as registered:
   Liv-ex approves vintages one by one. `vintage_config` and the first/final
   vintage columns are reported so the caller can hedge, and "NV" is deliberately
   not synthesised - the file is LWIN7-level and does not carry that convention,
   so inventing one here would be a guess wearing a code's clothes.

5. The base is a **snapshot**. A wine that is absent may simply postdate the
   file; `vin_base` reports the source's date so "not found" can be said with the
   right amount of doubt.

Storage is a SQLite mirror rebuilt whenever the source file's size or mtime
moves, kept outside the source directory (the volume may be read-only) and built
lazily on first use, in a thread: 200 000 rows must not be parsed at import time
in a hub that hosts a dozen other addons.
"""

from __future__ import annotations

import asyncio
import csv
import hashlib
import logging
import os
import re
import sqlite3
import tempfile
import time
import unicodedata
from pathlib import Path

from ._common import new_server

logger = logging.getLogger("rosetta.vin")

mcp = new_server("vin")

# Absent, the addon mounts `degraded` and says so on /health: the base is a
# manual deposit, and a silent empty cellar would be worse than a visible gap.
required_env = ["LWIN_DB"]

DOWNLOAD_URL = "https://www.liv-ex.com/lwin/"

# One LIKE scan over the whole table is ~100 ms at LWIN's size, which is fine;
# what is not fine is handing back an arbitrary slice of 40 000 "chateau" hits
# as if it were an answer. Past this, the count is reported and the caller is
# told to narrow down.
MAX_MATCHES = 400
MAX_RESULTS = 25


# -- text folding -----------------------------------------------------------

# NFKD decomposes é and ç; it leaves ø, ß, æ, œ, đ, ł exactly as they are. The
# LWIN guide names ø and ß among the characters the base does not carry, so the
# gap is precisely where French/German/Nordic labels live. Trap 1.
_UNDECOMPOSED = str.maketrans({
    "ø": "o", "Ø": "o", "ß": "ss", "æ": "ae", "Æ": "ae", "œ": "oe", "Œ": "oe",
    "đ": "d", "Đ": "d", "ð": "d", "Ð": "d", "ł": "l", "Ł": "l", "þ": "th",
    "Þ": "th", "ı": "i", "'": " ", "’": " ", "-": " ",
})


def fold(value) -> str:
    """Lowercase, accent-stripped, punctuation-flattened text for matching.

    Applied to BOTH sides: the base has no diacritics by design, the user has
    nothing but. "Château Léoville" and "Chateau Leoville" must meet somewhere,
    and this is where.
    """
    text = str(value or "").translate(_UNDECOMPOSED)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return text.strip()


# -- the source file --------------------------------------------------------

# Header -> canonical field. The key is the header folded down to letters, so
# "DISPLAY_NAME", "Display Name" and "displayName" all land on the same field.
COLUMNS = {
    "lwin": "lwin", "lwin7": "lwin", "lwinno": "lwin", "lwincode": "lwin",
    "displayname": "display_name", "display": "display_name",
    "producertitle": "producer_title", "producername": "producer_name",
    "producer": "producer_name",
    "wine": "wine", "winename": "wine",
    "country": "country", "region": "region", "subregion": "sub_region",
    "site": "site", "parcel": "parcel",
    "colour": "colour", "color": "colour",
    "type": "type", "subtype": "sub_type",
    "designation": "designation", "classification": "classification",
    "vintageconfig": "vintage_config", "vintageconfiguration": "vintage_config",
    "firstvintage": "first_vintage", "finalvintage": "final_vintage",
    "status": "status", "reference": "reference",
    "dateadded": "date_added", "dateupdated": "date_updated",
}

FIELDS = ["lwin", "display_name", "producer_title", "producer_name", "wine",
          "country", "region", "sub_region", "site", "parcel", "colour",
          "type", "sub_type", "designation", "classification",
          "vintage_config", "first_vintage", "final_vintage", "status"]

# What a search actually looks through. Deliberately not `status` or the dates:
# a query is a wine's name, not its bookkeeping.
SEARCHABLE = ["display_name", "producer_title", "producer_name", "wine",
              "country", "region", "sub_region", "site", "parcel",
              "classification", "designation", "colour", "type", "sub_type"]

_CODE = re.compile(r"^\d+$")


def _header_map(header: list[str]) -> dict[int, str]:
    """Column index -> canonical field, for the columns we recognise."""
    out = {}
    for i, raw in enumerate(header or []):
        key = re.sub(r"[^a-z0-9]+", "", str(raw or "").lower())
        if key in COLUMNS and COLUMNS[key] not in out.values():
            out[i] = COLUMNS[key]
    return out


def _normalise_lwin(raw) -> str | None:
    """A 7-digit LWIN, or None.

    Zero-padding is not politeness: a spreadsheet that has ever treated the
    column as a number drops the leading zero of `0123456`, and every code in
    the file comes out six digits long. Longer than 7 means the file is not
    LWIN7-level - counted as rejected rather than truncated into a wrong key.
    """
    code = re.sub(r"\D", "", str(raw or ""))
    if not code or len(code) > 7:
        return None
    return code.zfill(7)


def _rows_from_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(fh, dialect)
        for row in reader:
            yield row


def _rows_from_xlsx(path: Path):
    # Soft import: the LWIN download is an Excel workbook, but an addon must
    # never take the hub down over a missing wheel - the tools say it instead.
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise RuntimeError(
            "lecture .xlsx impossible : openpyxl n'est pas installé. "
            "Convertir la base LWIN en CSV, ou installer openpyxl."
        ) from exc
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        for row in sheet.iter_rows(values_only=True):
            yield ["" if c is None else str(c) for c in row]
    finally:
        book.close()


def read_rows(path: Path):
    """(header_map, iterator of canonical dicts) for a CSV or XLSX source."""
    reader = _rows_from_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") \
        else _rows_from_csv(path)
    columns: dict[int, str] = {}
    for row in reader:
        if not columns:
            columns = _header_map(row)
            if "lwin" not in columns.values():
                raise RuntimeError(
                    "en-tête illisible : aucune colonne LWIN reconnue "
                    f"dans {row[:8]}. Attendu le fichier « LWIN database » de Liv-ex."
                )
            continue
        record = {f: "" for f in FIELDS}
        for i, field in columns.items():
            if i < len(row):
                record[field] = str(row[i] or "").strip()
        yield record


# -- the SQLite mirror ------------------------------------------------------

SCHEMA = f"""
CREATE TABLE vins ({', '.join(f'{f} TEXT' for f in FIELDS)},
                   recherche TEXT, pays_pli TEXT, couleur_pli TEXT);
CREATE UNIQUE INDEX vins_lwin ON vins (lwin);
CREATE INDEX vins_pays ON vins (pays_pli);
CREATE TABLE meta (cle TEXT PRIMARY KEY, valeur TEXT);
"""


def source_path() -> Path | None:
    # Read per call, never captured at import: pointing the addon at a fresher
    # file is then a rollout, not a rebuild.
    raw = os.environ.get("LWIN_DB", "").strip()
    return Path(raw) if raw else None


def cache_path(source: Path) -> Path:
    """Where the mirror lives. Never beside the source: the volume carrying a
    manually deposited file is routinely mounted read-only."""
    override = os.environ.get("LWIN_CACHE", "").strip()
    if override:
        return Path(override)
    digest = hashlib.sha1(str(source.resolve()).encode()).hexdigest()[:12]
    return Path(tempfile.gettempdir()) / f"rosetta-lwin-{digest}.sqlite"


def _stamp(source: Path) -> str:
    st = source.stat()
    return f"{st.st_size}:{int(st.st_mtime)}"


def build(source: Path, cache: Path) -> dict:
    """Parse the source into a fresh SQLite mirror. Blocking - call in a thread."""
    started = time.monotonic()
    tmp = cache.with_suffix(cache.suffix + ".tmp")
    tmp.unlink(missing_ok=True)
    cache.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(tmp)
    try:
        conn.executescript(SCHEMA)
        placeholders = ", ".join("?" * (len(FIELDS) + 3))
        insert = f"INSERT OR IGNORE INTO vins VALUES ({placeholders})"
        kept = rejected = 0
        batch = []
        for record in read_rows(source):
            code = _normalise_lwin(record.get("lwin"))
            if not code:
                rejected += 1
                continue
            record["lwin"] = code
            # One folded haystack per row: the search is a LIKE over this, so
            # matching costs a scan and nothing else has to be normalised twice.
            haystack = fold(" ".join(record.get(f, "") for f in SEARCHABLE))
            batch.append([record[f] for f in FIELDS] +
                         [f" {haystack} ", fold(record.get("country")),
                          fold(record.get("colour"))])
            kept += 1
            if len(batch) >= 2000:
                conn.executemany(insert, batch)
                batch = []
        if batch:
            conn.executemany(insert, batch)
        meta = {
            "source": str(source),
            "empreinte": _stamp(source),
            "lignes": str(kept),
            "rejetees": str(rejected),
            "construite": time.strftime("%Y-%m-%d %H:%M:%S"),
            "source_datee": time.strftime("%Y-%m-%d",
                                          time.localtime(source.stat().st_mtime)),
        }
        conn.executemany("INSERT INTO meta VALUES (?, ?)", list(meta.items()))
        conn.commit()
    finally:
        conn.close()
    tmp.replace(cache)
    logger.info("base LWIN construite : %s lignes, %s rejetées, %.1fs",
                meta["lignes"], meta["rejetees"], time.monotonic() - started)
    return meta


def _meta(conn: sqlite3.Connection) -> dict:
    return dict(conn.execute("SELECT cle, valeur FROM meta").fetchall())


def _connect(cache: Path) -> sqlite3.Connection:
    # `check_same_thread=False` is required, not lax: the mirror is opened in one
    # `to_thread` worker and queried from another, and asyncio gives no promise
    # that it is the same one. Each tool call owns its connection and closes it,
    # so nothing is ever shared concurrently.
    conn = sqlite3.connect(cache, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def _open(source: Path, cache: Path) -> sqlite3.Connection:
    """A connection on an up-to-date mirror, rebuilding it if the source moved.
    Blocking - call in a thread."""
    if cache.exists():
        conn = _connect(cache)
        try:
            if _meta(conn).get("empreinte") == _stamp(source):
                return conn
        except sqlite3.DatabaseError:
            pass  # truncated or foreign file: rebuild over it
        conn.close()
    build(source, cache)
    return _connect(cache)


_lock = asyncio.Lock()


async def connect() -> tuple[sqlite3.Connection | None, dict | str]:
    """(connection, meta) or (None, explicit French message)."""
    source = source_path()
    if source is None:
        return None, ("base LWIN non configurée : poser le fichier téléchargé sur "
                      f"{DOWNLOAD_URL} (gratuit, licence Creative Commons) et "
                      "pointer LWIN_DB dessus.")
    if not source.exists():
        return None, (f"base LWIN introuvable : LWIN_DB pointe sur {source}, "
                      "qui n'existe pas. Le fichier est un dépôt manuel — "
                      f"le télécharger sur {DOWNLOAD_URL}.")
    # The lock spans the build so two calls arriving together parse the file
    # once, not twice; the reads that follow are short enough to serialise.
    async with _lock:
        try:
            conn = await asyncio.to_thread(_open, source, cache_path(source))
        except Exception as exc:
            logger.exception("base LWIN illisible")
            return None, f"base LWIN illisible ({source}) : {exc}"
    return conn, _meta(conn)


# -- shaping ----------------------------------------------------------------

COLOUR_FR = {"red": "rouge", "white": "blanc", "rose": "rosé",
             "mixed": "assortiment", "not applicable": None}

TYPE_FR = {"wine": "vin", "fortified wine": "vin muté", "spirit": "spiritueux",
           "sake": "saké", "vermouth": "vermouth", "other": "autre",
           "beer": "bière", "cider": "cidre"}

CONFIG_FR = {
    "sequential": "millésimé chaque année ou presque",
    "nonsequential": "millésimé certaines années seulement",
    "singlevintageonly": "un seul millésime connu, ou produit non millésimé",
}

# Bottle formats, in millilitres. Above the magnum the naming SPLITS by region -
# 3 litres is a double-magnum in Bordeaux and a jéroboam in Burgundy and
# Champagne - so both are given rather than one picked and passed off as the
# name.
FORMATS = {
    187: "quart", 200: "quart", 250: "25 cl", 375: "demi-bouteille",
    500: "50 cl", 620: "clavelin (vin jaune)", 750: "bouteille",
    1000: "litre", 1500: "magnum",
    3000: "double-magnum (Bordeaux) / jéroboam (Bourgogne, Champagne)",
    4500: "jéroboam (Bordeaux) / réhoboam (Bourgogne, Champagne)",
    6000: "impériale (Bordeaux) / mathusalem (Bourgogne, Champagne)",
    9000: "salmanazar", 12000: "balthazar", 15000: "nabuchodonosor",
}


def _put(out: dict, key: str, value) -> None:
    """A hole stays a hole: never hand the agent a blank it could read as data."""
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ("", "not applicable", "n/a", "none", "null", "unknown"):
            return
    if value in (None, "", [], {}):
        return
    out[key] = value


def shape(row) -> dict:
    """One LWIN7 record, projected onto what an agent (and a cellar) can use."""
    out: dict = {"lwin7": row["lwin"]}
    _put(out, "nom", row["display_name"])
    producer = " ".join(p for p in (row["producer_title"], row["producer_name"]) if p)
    _put(out, "producteur", producer)
    _put(out, "cuvee", row["wine"])
    _put(out, "couleur", COLOUR_FR.get((row["colour"] or "").lower().strip(),
                                       row["colour"]))
    _put(out, "type", TYPE_FR.get((row["type"] or "").lower().strip(), row["type"]))
    _put(out, "sous_type", row["sub_type"])
    _put(out, "pays", row["country"])
    _put(out, "region", row["region"])
    _put(out, "sous_region", row["sub_region"])
    _put(out, "lieu", row["site"])
    _put(out, "parcelle", row["parcel"])
    _put(out, "classification", row["classification"])
    _put(out, "appellation", row["designation"])
    config = re.sub(r"[^a-z]", "", (row["vintage_config"] or "").lower())
    _put(out, "millesimage", CONFIG_FR.get(config, row["vintage_config"]))
    _put(out, "premier_millesime", row["first_vintage"])
    _put(out, "dernier_millesime", row["final_vintage"])
    _put(out, "statut", row["status"])
    return out


def _year(value) -> int | None:
    digits = re.sub(r"\D", "", str(value or ""))
    return int(digits) if len(digits) == 4 else None


def vintage_check(row, year: int) -> str | None:
    """Why a vintage looks doubtful for this wine, or None. A caution, never a
    refusal: the file is a snapshot and Liv-ex is the one who arbitrates."""
    first, final = _year(row["first_vintage"]), _year(row["final_vintage"])
    config = re.sub(r"[^a-z]", "", (row["vintage_config"] or "").lower())
    if first and year < first:
        return f"antérieur au premier millésime connu ({first})"
    if final and year > final:
        return f"postérieur au dernier millésime connu ({final})"
    if config == "singlevintageonly":
        return ("ce LWIN7 n'a qu'un millésime connu, ou n'est pas millésimé — "
                "vérifier chez Liv-ex avant de composer un LWIN11")
    return None


# -- tools ------------------------------------------------------------------

@mcp.tool()
async def vin_recherche(query: str, pays: str | None = None,
                        couleur: str | None = None, millesime: int | None = None,
                        max_results: int = 8) -> dict:
    """Retrouver un vin dans la base LWIN (référentiel Liv-ex, hors ligne).

    query : le vin tel qu'on le lit sur l'étiquette — « Château Léoville Barton »,
            « Prum Wehlener Sonnenuhr », « Muga Prado Enea ». Les accents sont
            inutiles mais sans effet : la base n'en porte aucun, la recherche
            les replie des deux côtés.
    pays / couleur : filtres facultatifs, en anglais comme dans la base
            (« France », « Red »).
    millesime : facultatif. Compose le LWIN11 de chaque résultat et signale un
            millésime douteux au vu des premier/dernier millésimes connus.
    max_results : 8 par défaut, 25 au maximum.

    Rend, par vin : `lwin7`, nom d'affichage, producteur, cuvée, couleur, type,
    pays/région/sous-région/lieu, classification, appellation, mode de
    millésimage et millésimes connus.

    ⚠️ La base est un INSTANTANÉ téléchargé, pas un service : un vin absent peut
    simplement être postérieur au fichier (`vin_base` en donne la date). Et un
    LWIN7 identifie le vin, jamais la bouteille — le millésime, le contenant et
    le nombre de bouteilles s'ajoutent avec `vin_code`.
    """
    conn, meta = await connect()
    if conn is None:
        return {"error": meta}
    try:
        tokens = [t for t in fold(query).split() if t]
        if not tokens:
            return {"error": "recherche vide."}
        max_results = max(1, min(int(max_results or 8), MAX_RESULTS))

        where = ["recherche LIKE ?" for _ in tokens]
        params: list = [f"%{t}%" for t in tokens]
        if pays:
            where.append("pays_pli = ?")
            params.append(fold(pays))
        if couleur:
            where.append("couleur_pli = ?")
            params.append(fold(couleur))
        clause = " AND ".join(where)

        def query_db():
            total = conn.execute(
                f"SELECT COUNT(*) FROM vins WHERE {clause}", params).fetchone()[0]
            rows = conn.execute(
                f"SELECT * FROM vins WHERE {clause} LIMIT ?",
                params + [MAX_MATCHES]).fetchall()
            return total, rows

        total, rows = await asyncio.to_thread(query_db)

        def score(row) -> tuple:
            name = " " + fold(row["display_name"]) + " "
            # Three tiers, because two of them are not the same claim: "barton"
            # is a whole word in "Leoville Barton", only the start of one in
            # "Bartonia Estate", and merely buried inside "Charbartonnay".
            points = sum(3 if f" {t} " in name else 2 if f" {t}" in name
                         else 1 if t in name else 0
                         for t in tokens)
            if all(t in name for t in tokens):
                points += 3
            if (row["status"] or "").strip().lower() == "live":
                points += 1
            return (-points, len(row["display_name"] or ""), row["lwin"])

        rows = sorted(rows, key=score)[:max_results]
        results = []
        for row in rows:
            item = shape(row)
            if millesime is not None:
                year = _year(millesime)
                if year is None:
                    return {"error": "millésime attendu sur 4 chiffres (ex. 2015)."}
                item["lwin11"] = f"{row['lwin']}{year}"
                _put(item, "millesime_douteux", vintage_check(row, year))
            results.append(item)

        out = {"recherche": query, "resultats": results,
               "base_datee": meta.get("source_datee")}
        if total > len(rows):
            out["total"] = total
        if total > MAX_MATCHES:
            # An arbitrary slice of 40 000 hits reported as "the results" would
            # read as an answer. Say it instead.
            out["tronque"] = True
            out["note"] = (f"{total} vins correspondent — seuls {MAX_MATCHES} ont été "
                           "examinés. Préciser la recherche (producteur + cuvée, "
                           "ou ajouter `pays`).")
        return out
    finally:
        conn.close()


@mcp.tool()
async def vin_lwin(code: str) -> dict:
    """Décoder un code LWIN (7, 11, 16 ou 18 chiffres) et le résoudre.

    code : « 1012361 » (le vin), « 10123612009 » (+ millésime),
           « 1012361200900750 » (+ contenant), « 101236120091200750 »
           (+ nombre de bouteilles).

    Rend la fiche du vin (comme `vin_recherche`) plus ce que le code porte en
    propre : millésime, contenant en millilitres avec son nom français, nombre
    de bouteilles.

    ⚠️ Le nombre de bouteilles s'insère AU MILIEU, entre le millésime et le
    contenant : un LWIN18 n'est pas un LWIN16 suivi de deux chiffres. Et le
    contenant tient sur 5 chiffres avec ses zéros de tête (`00750`) — un code
    passé par un entier a perdu les siens et désigne autre chose.
    """
    digits = re.sub(r"\D", "", str(code or ""))
    if not digits:
        return {"error": "aucun code LWIN fourni."}
    if len(digits) not in (7, 11, 16, 18):
        return {"error": f"« {code} » fait {len(digits)} chiffres : un LWIN en "
                         "compte 7 (le vin), 11 (+ millésime), 16 (+ contenant) "
                         "ou 18 (+ nombre de bouteilles)."}

    lwin7, out = digits[:7], {"code": digits, "lwin7": digits[:7]}
    if len(digits) >= 11:
        out["millesime"] = digits[7:11]
    if len(digits) == 16:
        contenant = int(digits[11:16])
    elif len(digits) == 18:
        out["bouteilles"] = int(digits[11:13])
        contenant = int(digits[13:18])
    else:
        contenant = None
    if contenant is not None:
        out["contenant_ml"] = contenant
        _put(out, "format", FORMATS.get(contenant))

    conn, meta = await connect()
    if conn is None:
        out["note"] = f"code décodé, mais {meta}"
        return out
    try:
        row = await asyncio.to_thread(
            lambda: conn.execute("SELECT * FROM vins WHERE lwin = ?",
                                 (lwin7,)).fetchone())
        if row is None:
            out["trouve"] = False
            out["detail"] = (f"LWIN7 {lwin7} absent de la base du "
                             f"{meta.get('source_datee')} — le code reste "
                             "structurellement valide, il peut être plus récent "
                             "que le fichier.")
            return out
        out["trouve"] = True
        out.update(shape(row))
        year = _year(out.get("millesime"))
        if year is not None:
            _put(out, "millesime_douteux", vintage_check(row, year))
        return out
    finally:
        conn.close()


@mcp.tool()
async def vin_code(lwin7: str, millesime: int | None = None,
                   contenant_ml: int | None = None,
                   bouteilles: int | None = None) -> dict:
    """Composer le code LWIN d'une bouteille précise, pour l'inventaire d'une cave.

    lwin7 : le code du vin (7 chiffres), obtenu par `vin_recherche`.
    millesime : l'année sur 4 chiffres. Donne le LWIN11.
    contenant_ml : 750 pour une bouteille, 1500 pour un magnum… Avec le
            millésime, donne le LWIN16.
    bouteilles : nombre de bouteilles du lot (caisse de 6, de 12…). Avec les
            deux précédents, donne le LWIN18.

    Rend les codes composables, la fiche du vin, et une alerte si le millésime
    détonne avec ce que la base sait du vin.

    ⚠️ Un code composé ici est structurellement juste ; il n'atteste PAS que
    Liv-ex a enregistré ce millésime — leur validation se fait un à un. Pour un
    vin non millésimé, ne rien inventer : la base est au niveau du vin et ne
    porte pas la convention « NV ».
    """
    code = re.sub(r"\D", "", str(lwin7 or ""))
    if not code or len(code) > 7:
        return {"error": f"« {lwin7} » n'est pas un LWIN7 (7 chiffres attendus)."}
    code = code.zfill(7)

    year = _year(millesime) if millesime is not None else None
    if millesime is not None and year is None:
        return {"error": "millésime attendu sur 4 chiffres (ex. 2015)."}
    if contenant_ml is not None:
        contenant_ml = int(contenant_ml)
        if not 1 <= contenant_ml <= 99999:
            return {"error": "contenant attendu en millilitres, de 1 à 99999 "
                             "(750 pour une bouteille, 1500 pour un magnum)."}
        if year is None:
            return {"error": "un LWIN16 porte le millésime : préciser `millesime`."}
    if bouteilles is not None:
        bouteilles = int(bouteilles)
        if not 1 <= bouteilles <= 99:
            return {"error": "nombre de bouteilles attendu de 1 à 99 — le code "
                             "ne lui réserve que deux chiffres."}
        if contenant_ml is None:
            return {"error": "un LWIN18 porte le contenant : préciser "
                             "`contenant_ml` (750, 1500…)."}

    out: dict = {"lwin7": code}
    if year is not None:
        out["lwin11"] = f"{code}{year}"
        out["millesime"] = str(year)
    if contenant_ml is not None:
        # Trap 3: five digits, zero-padded, in millilitres.
        taille = f"{contenant_ml:05d}"
        out["lwin16"] = f"{code}{year}{taille}"
        out["contenant_ml"] = contenant_ml
        _put(out, "format", FORMATS.get(contenant_ml))
    if bouteilles is not None:
        # Trap 2: the pack size goes BETWEEN vintage and size, not at the end.
        out["lwin18"] = f"{code}{year}{bouteilles:02d}{contenant_ml:05d}"
        out["bouteilles"] = bouteilles

    conn, meta = await connect()
    if conn is None:
        out["note"] = f"codes composés, mais {meta}"
        return out
    try:
        row = await asyncio.to_thread(
            lambda: conn.execute("SELECT * FROM vins WHERE lwin = ?",
                                 (code,)).fetchone())
        if row is None:
            out["connu"] = False
            out["detail"] = (f"LWIN7 {code} absent de la base du "
                             f"{meta.get('source_datee')} : les codes ci-dessus "
                             "sont bien formés, mais rien ne confirme le vin.")
            return out
        out["connu"] = True
        out.update(shape(row))
        if year is not None:
            _put(out, "millesime_douteux", vintage_check(row, year))
        return out
    finally:
        conn.close()


@mcp.tool()
async def vin_base() -> dict:
    """État de la base LWIN locale : d'où elle vient, de quand elle date.

    À regarder avant de conclure qu'un vin « n'existe pas » : la base est un
    fichier téléchargé, donc un instantané. Rend le chemin source, sa date, le
    nombre de vins indexés et le nombre de lignes écartées à la construction.
    """
    conn, meta = await connect()
    if conn is None:
        return {"prete": False, "detail": meta, "telechargement": DOWNLOAD_URL}
    try:
        return {
            "prete": True,
            "source": meta.get("source"),
            "source_datee": meta.get("source_datee"),
            "vins": int(meta.get("lignes", 0)),
            "lignes_ecartees": int(meta.get("rejetees", 0)),
            "miroir_construit": meta.get("construite"),
            "telechargement": DOWNLOAD_URL,
        }
    finally:
        conn.close()


if __name__ == "__main__":
    # Local stdio debugging: `python -m rosetta.addons.vin`.
    mcp.run()
